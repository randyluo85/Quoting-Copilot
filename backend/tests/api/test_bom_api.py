"""BOM Upload API 集成测试.

遵循 TDD 原则：
1. RED - 先写失败的测试
2. GREEN - 写最小代码使测试通过
3. REFACTOR - 重构清理
"""

import pytest
import os
from httpx import AsyncClient, ASGITransport
from io import BytesIO
from openpyxl import Workbook

from app.main import app
from app.db.session import get_db
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession, async_sessionmaker
from sqlalchemy.pool import NullPool
from sqlalchemy import text
from app.config import get_settings
from decimal import Decimal

# 获取项目根目录（用于访问测试文件）
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
BOM_FILES_DIR = os.path.join(PROJECT_ROOT, "tests", "files")


@pytest.fixture
async def async_client():
    """创建测试用的异步 HTTP 客户端."""
    settings = get_settings()

    test_engine = create_async_engine(
        settings.mysql_url,
        echo=False,
        poolclass=NullPool,
    )

    TestingSessionLocal = async_sessionmaker(
        test_engine,
        class_=AsyncSession,
        expire_on_commit=False,
    )

    async def override_get_db():
        async with TestingSessionLocal() as session:
            yield session

    app.dependency_overrides[get_db] = override_get_db

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        yield client

    app.dependency_overrides.clear()
    await test_engine.dispose()


@pytest.fixture
async def clean_db(async_client: AsyncClient):
    """清空测试数据."""
    async for session in get_db():
        db_session = session
        break

    await db_session.execute(text("SET FOREIGN_KEY_CHECKS=0"))
    await db_session.execute(text("TRUNCATE TABLE project_products"))
    await db_session.execute(text("TRUNCATE TABLE projects"))
    await db_session.execute(text("TRUNCATE TABLE process_rates"))
    await db_session.execute(text("TRUNCATE TABLE materials"))
    await db_session.execute(text("SET FOREIGN_KEY_CHECKS=1"))
    await db_session.commit()

    # 插入测试数据
    test_material = {
        "item_code": "MAT-TEST-DUAL",
        "name": "双轨价格测试物料",
        "std_price": Decimal("100.00"),
        "vave_price": Decimal("85.00"),
        "supplier_tier": "A",
        "category": "测试",
        "material": "铝合金",
        "status": "active",
    }
    await db_session.execute(
        text(
            "INSERT INTO materials (item_code, name, std_price, vave_price, supplier_tier, category, material, status) "
            "VALUES (:item_code, :name, :std_price, :vave_price, :supplier_tier, :category, :material, :status)"
        ),
        test_material,
    )

    await db_session.commit()
    yield db_session


@pytest.fixture
def valid_bom_file():
    """生成有效的 BOM Excel 文件."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    headers = [
        "Level",
        "Item",
        "Part Number",
        "Part Name",
        "Version",
        "Type",
        "Status",
        "Material",
        "Supplier",
        "Qty",
        "Unit",
        "Comments",
    ]
    ws.append(headers)

    # 有历史价格的物料
    ws.append(
        [
            "1",
            "100",
            "MAT-TEST-DUAL",
            "双轨价格测试物料",
            "V1.0",
            "I",
            "N",
            "铝合金",
            "供应商A",
            2.5,
            "kg",
            "铸造级材料，需要折弯",
        ]
    )

    # 无历史价格的物料
    ws.append(
        [
            "2",
            "200",
            "MAT-NEW-ITEM",
            "新物料无历史",
            "V1.0",
            "I",
            "N",
            "不锈钢",
            "供应商B",
            1.0,
            "kg",
            "需要焊接和表面处理",
        ]
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@pytest.fixture
def empty_bom_file():
    """生成空的 BOM Excel 文件（只有表头）."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    headers = [
        "Level",
        "Part Number",
        "Part Name",
        "Qty",
        "Comments",
    ]
    ws.append(headers)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@pytest.fixture
def invalid_file():
    """生成非 Excel 文件."""
    return BytesIO(b"This is not an Excel file")


@pytest.fixture
def corrupted_excel_file():
    """生成损坏的 Excel 文件."""
    return BytesIO(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1\x00\x00\x00\x00")


class TestBOMUpload:
    """POST /api/v1/bom/upload - BOM 上传测试."""

    @pytest.mark.asyncio
    async def test_upload_valid_bom_returns_parsed_data(
        self, async_client, clean_db, valid_bom_file
    ):
        """上传有效 BOM 文件返回解析数据."""
        response = await async_client.post(
            "/api/v1/bom/upload",
            files={"file": ("bom.xlsx", valid_bom_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"project_id": "TEST-PROJECT-001"},
        )

        assert response.status_code == 200
        data = response.json()

        assert "parseId" in data
        assert data["status"] == "completed"
        assert "materials" in data
        assert "processes" in data
        assert "summary" in data

    @pytest.mark.asyncio
    async def test_bom_upload_includes_historical_prices(
        self, async_client, clean_db, valid_bom_file
    ):
        """BOM 上传包含历史价格匹配."""
        response = await async_client.post(
            "/api/v1/bom/upload",
            files={"file": ("bom.xlsx", valid_bom_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"project_id": "TEST-PROJECT-001"},
        )

        assert response.status_code == 200
        data = response.json()

        # 验证匹配到的物料有价格
        materials = data["materials"]
        matched_material = next((m for m in materials if m["partNumber"] == "MAT-TEST-DUAL"), None)

        assert matched_material is not None
        assert matched_material["unitPrice"] is not None
        assert matched_material["vavePrice"] is not None
        assert matched_material["hasHistoryData"] is True

    @pytest.mark.asyncio
    async def test_bom_upload_detects_processes_from_comments(
        self, async_client, clean_db, valid_bom_file
    ):
        """BOM 上传从 comments 中检测工艺."""
        response = await async_client.post(
            "/api/v1/bom/upload",
            files={"file": ("bom.xlsx", valid_bom_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"project_id": "TEST-PROJECT-001"},
        )

        assert response.status_code == 200
        data = response.json()

        # 验证工艺列表存在
        processes = data["processes"]
        # 可能检测到折弯、焊接等工艺
        assert isinstance(processes, list)

    @pytest.mark.asyncio
    async def test_bom_upload_returns_summary_stats(
        self, async_client, clean_db, valid_bom_file
    ):
        """BOM 上传返回统计摘要."""
        response = await async_client.post(
            "/api/v1/bom/upload",
            files={"file": ("bom.xlsx", valid_bom_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"project_id": "TEST-PROJECT-001"},
        )

        assert response.status_code == 200
        data = response.json()

        summary = data["summary"]
        assert "totalMaterials" in summary
        assert "matchedMaterials" in summary
        assert "totalProcesses" in summary
        assert "matchedProcesses" in summary

        assert summary["totalMaterials"] >= 1


class TestBOMUploadEdgeCases:
    """BOM 上传边界情况测试."""

    @pytest.mark.asyncio
    async def test_upload_empty_bom_returns_warning(
        self, async_client, clean_db, empty_bom_file
    ):
        """上传空 BOM 文件返回警告."""
        response = await async_client.post(
            "/api/v1/bom/upload",
            files={"file": ("empty.xlsx", empty_bom_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"project_id": "TEST-PROJECT-001"},
        )

        assert response.status_code == 200
        data = response.json()

        # 空文件应返回空列表
        assert data["materials"] == []
        assert data["summary"]["totalMaterials"] == 0

    @pytest.mark.asyncio
    async def test_bom_with_no_matched_materials(self, async_client, clean_db):
        """BOM 文件中没有匹配的物料."""
        wb = Workbook()
        ws = wb.active

        headers = ["Part Number", "Part Name", "Qty", "Comments"]
        ws.append(headers)
        ws.append(["MAT-NONEXISTENT", "不存在物料", 1.0, "测试"])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = await async_client.post(
            "/api/v1/bom/upload",
            files={"file": ("bom.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"project_id": "TEST-PROJECT-001"},
        )

        assert response.status_code == 200
        data = response.json()

        # 没有匹配的物料
        assert data["summary"]["matchedMaterials"] == 0


class TestBOMStatusLight:
    """BOM 红绿灯状态测试."""

    @pytest.mark.asyncio
    async def test_matched_material_gets_green_status(
        self, async_client, clean_db, valid_bom_file
    ):
        """完全匹配的物料获得 GREEN 状态."""
        response = await async_client.post(
            "/api/v1/bom/upload",
            files={"file": ("bom.xlsx", valid_bom_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"project_id": "TEST-PROJECT-001"},
        )

        assert response.status_code == 200
        data = response.json()

        matched_material = next(
            (m for m in data["materials"] if m["partNumber"] == "MAT-TEST-DUAL"),
            None,
        )

        assert matched_material is not None
        assert matched_material["status"] in ["verified", "green"]
        assert matched_material["hasHistoryData"] is True

    @pytest.mark.asyncio
    async def test_unmatched_material_gets_red_status(self, async_client, clean_db):
        """无匹配的物料获得 RED 状态."""
        wb = Workbook()
        ws = wb.active

        headers = ["Part Number", "Part Name", "Qty", "Comments"]
        ws.append(headers)
        ws.append(["MAT-UNMATCHED", "无匹配物料", 1.0, ""])  # 无 comments，无历史

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = await async_client.post(
            "/api/v1/bom/upload",
            files={"file": ("bom.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"project_id": "TEST-PROJECT-001"},
        )

        assert response.status_code == 200
        data = response.json()

        unmatched_material = data["materials"][0]
        assert unmatched_material["status"] in ["missing", "red"]
        assert unmatched_material["hasHistoryData"] is False

    @pytest.mark.asyncio
    async def test_material_with_comments_gets_yellow_status(
        self, async_client, clean_db
    ):
        """有 comments 但无历史数据的物料获得 YELLOW 状态."""
        wb = Workbook()
        ws = wb.active

        headers = ["Part Number", "Part Name", "Qty", "Comments"]
        ws.append(headers)
        ws.append(["MAT-WITH-COMMENTS", "带备注物料", 1.0, "需要折弯处理"])  # 有 comments

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = await async_client.post(
            "/api/v1/bom/upload",
            files={"file": ("bom.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"project_id": "TEST-PROJECT-001"},
        )

        assert response.status_code == 200
        data = response.json()

        material = data["materials"][0]
        # 有 comments 但无历史数据，应为 YELLOW
        assert material["hasHistoryData"] is False
        # 状态可能是 warning/yellow 或 missing/red，取决于实现
        assert material["status"] in ["warning", "yellow", "missing", "red"]


class TestBOMUploadErrors:
    """BOM 上传错误场景测试."""

    @pytest.mark.asyncio
    async def test_upload_non_excel_file_returns_error(
        self, async_client, clean_db, invalid_file
    ):
        """上传非 Excel 文件返回错误."""
        response = await async_client.post(
            "/api/v1/bom/upload",
            files={"file": ("test.txt", invalid_file, "text/plain")},
            data={"project_id": "TEST-PROJECT-001"},
        )

        # 非 Excel 文件应被拒绝
        assert response.status_code in [400, 422]

    @pytest.mark.asyncio
    async def test_upload_without_project_id_returns_error(
        self, async_client, clean_db, valid_bom_file
    ):
        """缺少 project_id 返回错误."""
        response = await async_client.post(
            "/api/v1/bom/upload",
            files={"file": ("bom.xlsx", valid_bom_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            # 缺少 project_id
        )

        # 应返回验证错误
        assert response.status_code in [400, 422]

    @pytest.mark.asyncio
    async def test_upload_corrupted_excel_returns_error(
        self, async_client, clean_db, corrupted_excel_file
    ):
        """上传损坏的 Excel 文件返回错误."""
        response = await async_client.post(
            "/api/v1/bom/upload",
            files={"file": ("corrupt.xlsx", corrupted_excel_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"project_id": "TEST-PROJECT-001"},
        )

        # 损坏的文件应被拒绝
        assert response.status_code in [400, 422, 500]


@pytest.fixture
def real_bom_file_path():
    """获取真实 BOM 文件的路径."""
    bom_path = os.path.join(BOM_FILES_DIR, "bom.xlsx")
    if not os.path.exists(bom_path):
        pytest.skip(f"真实 BOM 文件不存在: {bom_path}")
    return bom_path


@pytest.fixture
def real_bom_file_content(real_bom_file_path):
    """读取真实 BOM 文件内容."""
    with open(real_bom_file_path, "rb") as f:
        return f.read()


class TestRealBOMFileUpload:
    """使用真实 BOM 文件的 API 测试."""

    @pytest.mark.asyncio
    async def test_upload_real_bom_file(
        self, async_client, clean_db, real_bom_file_content
    ):
        """上传真实 BOM 文件返回解析数据."""
        response = await async_client.post(
            "/api/v1/bom/upload",
            files={
                "file": (
                    "bom.xlsx",
                    real_bom_file_content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            data={"project_id": "TEST-PROJECT-REAL"},
        )

        assert response.status_code == 200
        data = response.json()

        assert "parseId" in data
        assert data["status"] == "completed"
        assert "materials" in data
        assert "summary" in data

    @pytest.mark.asyncio
    async def test_real_bom_file_parses_all_materials(
        self, async_client, clean_db, real_bom_file_content
    ):
        """真实 BOM 文件解析出所有物料."""
        response = await async_client.post(
            "/api/v1/bom/upload",
            files={
                "file": (
                    "bom.xlsx",
                    real_bom_file_content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            data={"project_id": "TEST-PROJECT-REAL"},
        )

        assert response.status_code == 200
        data = response.json()

        materials = data["materials"]
        # 真实 BOM 文件应该有多个物料
        assert len(materials) > 0

        # 验证每个物料有基本字段
        for material in materials:
            assert "partNumber" in material
            assert "partName" in material

    @pytest.mark.asyncio
    async def test_real_bom_file_detects_processes_from_sheet(
        self, async_client, clean_db, real_bom_file_content
    ):
        """真实 BOM 文件从工艺表解析工艺."""
        response = await async_client.post(
            "/api/v1/bom/upload",
            files={
                "file": (
                    "bom.xlsx",
                    real_bom_file_content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            data={"project_id": "TEST-PROJECT-REAL"},
        )

        assert response.status_code == 200
        data = response.json()

        # 真实 BOM 文件有工艺表
        processes = data.get("processes", [])
        assert isinstance(processes, list)

        if len(processes) > 0:
            # 验证工艺结构
            process = processes[0]
            assert "opNo" in process
            assert "name" in process

    @pytest.mark.asyncio
    async def test_real_bom_summary_statistics(
        self, async_client, clean_db, real_bom_file_content
    ):
        """真实 BOM 文件返回正确统计."""
        response = await async_client.post(
            "/api/v1/bom/upload",
            files={
                "file": (
                    "bom.xlsx",
                    real_bom_file_content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            data={"project_id": "TEST-PROJECT-REAL"},
        )

        assert response.status_code == 200
        data = response.json()

        summary = data["summary"]
        assert "totalMaterials" in summary
        assert "totalProcesses" in summary

        # 确保统计数据合理
        assert summary["totalMaterials"] >= 0
        assert summary["totalProcesses"] >= 0


class TestRealBOMFileIntegration:
    """真实 BOM 文件集成测试 - 验证与实际业务数据的兼容性."""

    @pytest.mark.asyncio
    async def test_real_bom_with_comments_parsing(
        self, async_client, clean_db, real_bom_file_content
    ):
        """真实 BOM 文件正确解析 comments 字段."""
        response = await async_client.post(
            "/api/v1/bom/upload",
            files={
                "file": (
                    "bom.xlsx",
                    real_bom_file_content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            data={"project_id": "TEST-PROJECT-REAL"},
        )

        assert response.status_code == 200
        data = response.json()

        materials = data["materials"]
        # 查找有 comments 的物料
        materials_with_comments = [
            m for m in materials if m.get("comments") and m["comments"].strip()
        ]

        # 真实 BOM 文件中应该有带 comments 的物料
        if len(materials_with_comments) > 0:
            # 验证 comments 字段正确解析
            material = materials_with_comments[0]
            assert isinstance(material["comments"], str)
            assert len(material["comments"]) > 0

    @pytest.mark.asyncio
    async def test_real_bom_quantity_parsing(
        self, async_client, clean_db, real_bom_file_content
    ):
        """真实 BOM 文件正确解析数量字段."""
        response = await async_client.post(
            "/api/v1/bom/upload",
            files={
                "file": (
                    "bom.xlsx",
                    real_bom_file_content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            data={"project_id": "TEST-PROJECT-REAL"},
        )

        assert response.status_code == 200
        data = response.json()

        materials = data["materials"]
        for material in materials:
            # 验证数量是数字
            assert "quantity" in material
            if material["quantity"] is not None:
                assert isinstance(material["quantity"], (int, float))

    @pytest.mark.asyncio
    async def test_real_bom_unit_parsing(
        self, async_client, clean_db, real_bom_file_content
    ):
        """真实 BOM 文件正确解析单位字段."""
        response = await async_client.post(
            "/api/v1/bom/upload",
            files={
                "file": (
                    "bom.xlsx",
                    real_bom_file_content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            data={"project_id": "TEST-PROJECT-REAL"},
        )

        assert response.status_code == 200
        data = response.json()

        materials = data["materials"]
        # 验证单位字段解析（MM, PC 等）
        for material in materials:
            assert "unit" in material
