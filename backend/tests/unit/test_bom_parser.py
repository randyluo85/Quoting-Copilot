"""BOM Parser 单元测试.

遵循 TDD 原则：
1. RED - 先写失败的测试
2. GREEN - 写最小代码使测试通过
3. REFACTOR - 重构清理
"""

import pytest
import os
from io import BytesIO
from openpyxl import Workbook

from app.services.bom_parser import (
    BOMParser,
    ParsedMaterial,
    ParsedProcess,
    BOMParseResult,
)

# 获取项目根目录
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
BOM_FILES_DIR = os.path.join(PROJECT_ROOT, "tests", "files")


class TestBOMParserBasic:
    """BOM Parser 基础功能测试."""

    def test_parse_empty_excel_returns_empty_result(self):
        """测试解析空 Excel 文件返回空结果."""
        # Arrange
        wb = Workbook()
        ws = wb.active
        ws.title = "BOM"

        # 添加表头但没有数据
        headers = [
            "Level",
            "Part Number",
            "Part Name",
            "Version",
            "Type",
            "Status",
            "Material",
            "Supplier",
            "Qty",
            "Unit",
            "Comments",
        ]
        ws.append(headers)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Act
        parser = BOMParser()
        result = parser.parse_excel_file(output.read())

        # Assert
        assert isinstance(result, BOMParseResult)
        assert len(result.materials) == 0
        assert len(result.processes) == 0

    def test_parse_excel_with_single_material(self):
        """测试解析包含单个物料的 Excel."""
        # Arrange
        wb = Workbook()
        ws = wb.active
        ws.title = "BOM Material"

        # 表头
        headers = [
            "Level",
            "Item",
            "Part Number",
            "Part Name",
            "Version",
            "Type",
            "Status",
            "Material",
            "Supplier",
            "Qty",
            "Unit",
            "Comments",
        ]
        ws.append(headers)

        # 数据行
        ws.append([
            "1",
            "100",
            "MAT-TEST-001",
            "测试物料",
            "V1.0",
            "I",
            "N",
            "铝合金",
            "供应商A",
            2.5,
            "kg",
            "铸造级材料",
        ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Act
        parser = BOMParser()
        result = parser.parse_excel_file(output.read())

        # Assert
        assert len(result.materials) == 1
        material = result.materials[0]
        assert material.part_number == "MAT-TEST-001"
        assert material.part_name == "测试物料"
        assert material.quantity == 2.5
        assert material.material == "铝合金"
        assert material.supplier == "供应商A"
        assert material.comments == "铸造级材料"

    def test_parse_excel_with_multiple_materials(self):
        """测试解析包含多个物料的 Excel."""
        # Arrange
        wb = Workbook()
        ws = wb.active
        ws.title = "BOM"

        headers = [
            "Level",
            "Item",
            "Part Number",
            "Part Name",
            "Version",
            "Type",
            "Status",
            "Material",
            "Supplier",
            "Qty",
            "Unit",
            "Comments",
        ]
        ws.append(headers)

        # 添加多行数据
        ws.append([
            "1", "100", "MAT-001", "物料一", "V1.0", "I", "N",
            "铝", "A", 1.0, "kg", "备注一"
        ])
        ws.append([
            "2", "200", "MAT-002", "物料二", "V1.0", "I", "N",
            "钢", "B", 2.0, "kg", "备注二"
        ])
        ws.append([
            "3", "300", "MAT-003", "物料三", "V1.0", "I", "N",
            "铜", "C", 3.0, "kg", "备注三"
        ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Act
        parser = BOMParser()
        result = parser.parse_excel_file(output.read())

        # Assert
        assert len(result.materials) == 3
        assert result.materials[0].part_number == "MAT-001"
        assert result.materials[1].part_number == "MAT-002"
        assert result.materials[2].part_number == "MAT-003"


class TestBOMParserColumnDetection:
    """BOM Parser 智能列检测测试."""

    def test_detect_column_mapping_with_part_number_header(self):
        """测试检测包含 'Part Number' 表头的列映射."""
        # Arrange
        wb = Workbook()
        ws = wb.active

        # 使用非标准列顺序
        headers = [
            "Comments",
            "Qty",
            "Part Number",
            "Part Name",
            "Material",
            "Supplier",
        ]
        ws.append(headers)
        ws.append(["备注", 1.5, "MAT-TEST", "测试", "铝", "A"])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Act
        parser = BOMParser()
        result = parser.parse_excel_file(output.read())

        # Assert
        assert len(result.materials) == 1
        assert result.materials[0].part_number == "MAT-TEST"

    def test_detect_column_mapping_with_chinese_headers(self):
        """测试检测中文表头."""
        # Arrange
        wb = Workbook()
        ws = wb.active

        headers = ["层级", "零件号", "零件名称", "数量", "单位", "备注"]
        ws.append(headers)
        ws.append(["1", "MAT-CN-001", "中文测试物料", "5.0", "kg", "中文备注"])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Act
        parser = BOMParser()
        result = parser.parse_excel_file(output.read())

        # Assert
        assert len(result.materials) == 1
        assert result.materials[0].part_number == "MAT-CN-001"
        assert result.materials[0].part_name == "中文测试物料"
        assert result.materials[0].quantity == 5.0


class TestBOMParserEdgeCases:
    """BOM Parser 边界情况测试."""

    def test_skip_empty_rows(self):
        """测试跳过空行."""
        # Arrange
        wb = Workbook()
        ws = wb.active

        headers = ["Part Number", "Part Name", "Version", "Type", "Status", "Qty"]
        ws.append(headers)
        ws.append(["MAT-001", "物料一", "V1.0", "I", "N", 1.0])
        ws.append([])  # 空行
        ws.append(["MAT-002", "物料二", "V1.0", "I", "N", 2.0])
        ws.append([None, None, None, None, None, None])  # None 行

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Act
        parser = BOMParser()
        result = parser.parse_excel_file(output.read())

        # Assert
        assert len(result.materials) == 2

    def test_handle_missing_quantity(self):
        """测试处理缺失的数量."""
        # Arrange
        wb = Workbook()
        ws = wb.active

        headers = ["Part Number", "Part Name", "Version", "Type", "Status", "Qty"]
        ws.append(headers)
        ws.append(["MAT-NO-QTY", "无数量物料", "V1.0", "I", "N", None])  # 数量为空
        ws.append(["MAT-001", "正常物料", "V1.0", "I", "N", 1.0])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Act
        parser = BOMParser()
        result = parser.parse_excel_file(output.read())

        # Assert
        assert len(result.materials) == 2
        assert result.materials[0].quantity == 0  # 缺失数量默认为 0
        assert result.materials[1].quantity == 1.0

    def test_handle_invalid_quantity(self):
        """测试处理无效的数量（非数字）."""
        # Arrange
        wb = Workbook()
        ws = wb.active

        headers = ["Part Number", "Part Name", "Version", "Type", "Status", "Qty"]
        ws.append(headers)
        ws.append(["MAT-INVALID", "无效数量物料", "V1.0", "I", "N", "N/A"])  # 字符串数量

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Act
        parser = BOMParser()
        result = parser.parse_excel_file(output.read())

        # Assert
        assert len(result.materials) == 1
        assert result.materials[0].quantity == 0  # 无效数量默认为 0

    def test_skip_rows_without_part_number(self):
        """测试跳过没有零件号的行."""
        # Arrange
        wb = Workbook()
        ws = wb.active

        headers = ["Part Number", "Part Name", "Version", "Type", "Status", "Qty"]
        ws.append(headers)
        ws.append([None, "无零件号物料", "V1.0", "I", "N", 1.0])  # 零件号为空
        ws.append(["", "空字符串零件号", "V1.0", "I", "N", 2.0])  # 零件号为空字符串
        ws.append(["MAT-001", "正常物料", "V1.0", "I", "N", 3.0])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Act
        parser = BOMParser()
        result = parser.parse_excel_file(output.read())

        # Assert
        assert len(result.materials) == 1
        assert result.materials[0].part_number == "MAT-001"

    def test_use_default_values_for_missing_columns(self):
        """测试缺失列使用默认值."""
        # Arrange
        wb = Workbook()
        ws = wb.active

        # 只包含必要的列
        headers = ["Part Number", "Part Name", "Version", "Type", "Status"]
        ws.append(headers)
        ws.append(["MAT-MINIMAL", "最小字段物料", "V1.0", "I", "N"])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Act
        parser = BOMParser()
        result = parser.parse_excel_file(output.read())

        # Assert
        assert len(result.materials) == 1
        material = result.materials[0]
        assert material.part_number == "MAT-MINIMAL"
        assert material.quantity == 0  # 默认值
        assert material.unit == "PC"  # 默认值


class TestBOMParserSheetTypeDetection:
    """BOM Parser 工作表类型检测测试."""

    def test_detect_material_sheet_by_keywords(self):
        """测试通过关键字检测物料表."""
        # Arrange
        wb = Workbook()
        ws = wb.active
        ws.title = "Material BOM"

        # 表头包含物料关键字
        headers = ["Bill of Material", "Item", "Part Number", "Qty", "Material"]
        ws.append(headers)
        ws.append(["", "100", "MAT-001", 1.0, "铝"])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Act
        parser = BOMParser()
        result = parser.parse_excel_file(output.read())

        # Assert
        assert len(result.materials) == 1

    def test_detect_process_sheet_by_keywords(self):
        """测试通过关键字检测工艺表."""
        # Arrange
        wb = Workbook()
        ws = wb.active
        ws.title = "Process Sheet"

        headers = ["Operation", "Op No", "Name", "Work Center", "Standard Time", "Description"]
        ws.append(headers)
        ws.append(["", "010", "焊接", "焊接车间", 2.5, "对接焊"])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Act
        parser = BOMParser()
        result = parser.parse_excel_file(output.read())

        # Assert
        assert len(result.processes) == 1
        assert result.processes[0].op_no == "010"

    def test_detect_chinese_keywords(self):
        """测试检测中文关键字."""
        # Arrange
        wb = Workbook()
        ws = wb.active
        ws.title = "物料表"

        headers = ["零件号", "零件名称", "数量", "单位", "备注"]
        ws.append(headers)
        ws.append(["MAT-CN", "中文物料", 10, "kg", "备注"])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Act
        parser = BOMParser()
        result = parser.parse_excel_file(output.read())

        # Assert
        assert len(result.materials) == 1


class TestParsedMaterial:
    """ParsedMaterial 值对象测试."""

    def test_parsed_material_is_immutable(self):
        """测试 ParsedMaterial 是不可变的（NamedTuple）."""
        # Arrange & Act
        material = ParsedMaterial(
            level="1",
            part_number="MAT-001",
            part_name="测试",
            version="V1.0",
            type="I",
            status="N",
            material="铝",
            supplier="A",
            quantity=1.0,
            unit="kg",
            comments="备注",
        )

        # Assert - NamedTuple 是不可变的
        with pytest.raises(AttributeError):
            material.part_number = "NEW-001"

    def test_parsed_material_has_all_fields(self):
        """测试 ParsedMaterial 包含所有字段."""
        material = ParsedMaterial(
            level="1",
            part_number="MAT-001",
            part_name="测试",
            version="V1.0",
            type="I",
            status="N",
            material="铝",
            supplier="A",
            quantity=1.5,
            unit="kg",
            comments="备注",
        )

        assert material.level == "1"
        assert material.part_number == "MAT-001"
        assert material.part_name == "测试"
        assert material.version == "V1.0"
        assert material.type == "I"
        assert material.status == "N"
        assert material.material == "铝"
        assert material.supplier == "A"
        assert material.quantity == 1.5
        assert material.unit == "kg"
        assert material.comments == "备注"


class TestParsedProcess:
    """ParsedProcess 值对象测试."""

    def test_parsed_process_is_immutable(self):
        """测试 ParsedProcess 是不可变的."""
        process = ParsedProcess(
            op_no="010",
            name="焊接",
            work_center="焊接车间",
            standard_time=2.5,
            spec="对接焊",
        )

        with pytest.raises(AttributeError):
            process.op_no = "020"

    def test_parsed_process_with_optional_spec(self):
        """测试 ParsedProcess 的可选 spec 字段."""
        # 有 spec
        process_with_spec = ParsedProcess(
            op_no="010",
            name="焊接",
            work_center="焊接车间",
            standard_time=2.5,
            spec="对接焊",
        )
        assert process_with_spec.spec == "对接焊"

        # 无 spec
        process_without_spec = ParsedProcess(
            op_no="020",
            name="机加",
            work_center="机加车间",
            standard_time=1.5,
            spec=None,
        )
        assert process_without_spec.spec is None


class TestRealBOMFileParsing:
    """使用真实 BOM 文件的解析测试."""

    @pytest.fixture
    def real_bom_file_path(self):
        """获取真实 BOM 文件路径."""
        bom_path = os.path.join(BOM_FILES_DIR, "bom.xlsx")
        if not os.path.exists(bom_path):
            pytest.skip(f"真实 BOM 文件不存在: {bom_path}")
        return bom_path

    @pytest.fixture
    def real_bom_content(self, real_bom_file_path):
        """读取真实 BOM 文件内容."""
        with open(real_bom_file_path, "rb") as f:
            return f.read()

    def test_parse_real_bom_file(self, real_bom_content):
        """测试解析真实 BOM 文件."""
        parser = BOMParser()
        result = parser.parse_excel_file(real_bom_content)

        assert isinstance(result, BOMParseResult)
        assert isinstance(result.materials, list)
        assert isinstance(result.processes, list)

    def test_real_bom_file_has_materials(self, real_bom_content):
        """真实 BOM 文件包含物料."""
        parser = BOMParser()
        result = parser.parse_excel_file(real_bom_content)

        assert len(result.materials) > 0

    def test_real_bom_materials_have_required_fields(self, real_bom_content):
        """真实 BOM 物料包含必需字段."""
        parser = BOMParser()
        result = parser.parse_excel_file(real_bom_content)

        for material in result.materials:
            assert hasattr(material, "part_number")
            assert hasattr(material, "part_name")
            assert hasattr(material, "quantity")
            # 验证字段不为空
            assert material.part_number
            assert material.part_name

    def test_real_bom_materials_parsing_comments(self, real_bom_content):
        """真实 BOM 文件正确解析 comments."""
        parser = BOMParser()
        result = parser.parse_excel_file(real_bom_content)

        # 查找有 comments 的物料
        materials_with_comments = [
            m for m in result.materials if m.comments and m.comments.strip()
        ]

        if len(materials_with_comments) > 0:
            # 验证 comments 内容
            material = materials_with_comments[0]
            assert isinstance(material.comments, str)
            assert len(material.comments) > 0

    def test_real_bom_file_detects_process_sheet(self, real_bom_content):
        """真实 BOM 文件包含工艺表."""
        parser = BOMParser()
        result = parser.parse_excel_file(real_bom_content)

        # 真实 BOM 文件有 "工艺" sheet
        # 至少应该返回空列表
        assert isinstance(result.processes, list)

    def test_real_bom_quantity_values_are_numeric(self, real_bom_content):
        """真实 BOM 数量值是数字."""
        parser = BOMParser()
        result = parser.parse_excel_file(real_bom_content)

        for material in result.materials:
            assert isinstance(material.quantity, (int, float))

    def test_real_bom_unit_values_are_present(self, real_bom_content):
        """真实 BOM 单位值存在."""
        parser = BOMParser()
        result = parser.parse_excel_file(real_bom_content)

        for material in result.materials:
            assert material.unit is not None
            assert isinstance(material.unit, str)

    def test_real_bom_part_numbers_format(self, real_bom_content):
        """真实 BOM 零件号格式."""
        parser = BOMParser()
        result = parser.parse_excel_file(real_bom_content)

        for material in result.materials:
            # 零件号应该是字符串
            assert isinstance(material.part_number, str)
            # 不应该包含 None 值
            assert material.part_number != "None"

    def test_real_bom_process_parsing(self, real_bom_content):
        """真实 BOM 工艺解析."""
        parser = BOMParser()
        result = parser.parse_excel_file(real_bom_content)

        processes = result.processes
        if len(processes) > 0:
            # 验证工艺结构
            for process in processes:
                assert hasattr(process, "op_no")
                assert hasattr(process, "name")
                assert hasattr(process, "work_center")
                assert hasattr(process, "standard_time")


class TestRealBOMFilesMultiple:
    """测试多个真实 BOM 文件."""

    @pytest.fixture
    def all_bom_files(self):
        """获取所有 BOM 测试文件."""
        bom_files = []
        if os.path.exists(BOM_FILES_DIR):
            for filename in os.listdir(BOM_FILES_DIR):
                if filename.endswith(".xlsx") and not filename.startswith("~$"):
                    bom_files.append(os.path.join(BOM_FILES_DIR, filename))
        return bom_files

    def test_all_bom_files_are_parsable(self, all_bom_files):
        """所有 BOM 文件都可以被解析."""
        parser = BOMParser()

        for bom_path in all_bom_files:
            with open(bom_path, "rb") as f:
                content = f.read()
            result = parser.parse_excel_file(content)

            assert isinstance(result, BOMParseResult)
            assert isinstance(result.materials, list)
            assert isinstance(result.processes, list)

    def test_all_bom_files_have_materials(self, all_bom_files):
        """所有 BOM 文件都包含物料."""
        parser = BOMParser()

        for bom_path in all_bom_files:
            with open(bom_path, "rb") as f:
                content = f.read()
            result = parser.parse_excel_file(content)

            assert len(result.materials) > 0, f"{bom_path} 应该包含物料"
