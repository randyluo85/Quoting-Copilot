# backend/app/api/v1/materials.py
from __future__ import annotations

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from typing import Optional
from io import BytesIO
from decimal import Decimal

from app.schemas.material import (
    MaterialCreate,
    MaterialUpdate,
    MaterialResponse,
    MaterialListResponse,
    MaterialImportResult,
)
from app.services.material_service import MaterialService
from app.core.database import get_db

router = APIRouter(prefix="/materials", tags=["materials"])


@router.post("", response_model=MaterialResponse)
async def create_material(
    data: MaterialCreate,
    db: AsyncSession = Depends(get_db),
):
    """创建物料"""
    service = MaterialService(db)
    return await service.create(data)


@router.get("/{material_id}", response_model=MaterialResponse)
async def get_material(
    material_id: str,
    db: AsyncSession = Depends(get_db),
):
    """获取物料详情"""
    service = MaterialService(db)
    result = await service.get_by_id(material_id)
    if not result:
        raise HTTPException(status_code=404, detail="物料不存在")
    return result


@router.put("/{material_id}", response_model=MaterialResponse)
async def update_material(
    material_id: str,
    data: MaterialUpdate,
    db: AsyncSession = Depends(get_db),
):
    """更新物料"""
    service = MaterialService(db)
    result = await service.update(material_id, data)
    if not result:
        raise HTTPException(status_code=404, detail="物料不存在")
    return result


@router.delete("/{material_id}")
async def delete_material(
    material_id: str,
    db: AsyncSession = Depends(get_db),
):
    """删除物料"""
    service = MaterialService(db)
    success = await service.delete(material_id)
    if not success:
        raise HTTPException(status_code=404, detail="物料不存在")
    return {"message": "删除成功"}


@router.get("", response_model=MaterialListResponse)
async def list_materials(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """分页查询物料"""
    service = MaterialService(db)
    return await service.list(page, page_size, search)


@router.post("/import", response_model=MaterialImportResult)
async def import_materials(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Excel 批量导入物料"""
    from openpyxl import load_workbook

    service = MaterialService(db)

    # 读取 Excel
    contents = await file.read()
    workbook = load_workbook(filename=BytesIO(contents))
    sheet = workbook.active

    items = []
    errors = []

    # 从第 2 行开始（第 1 行是表头）
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # 假设 Excel 列顺序: id, name, spec, std_price, vave_price, supplier_tier
            item = MaterialCreate(
                id=str(row[0]),
                name=str(row[1]),
                spec=str(row[2]) if row[2] else None,
                std_price=Decimal(str(row[3])),
                vave_price=Decimal(str(row[4])) if row[4] else None,
                supplier_tier=str(row[5]) if row[5] else None,
            )
            items.append(item)
        except Exception as e:
            errors.append({"row": idx, "reason": str(e), "data": row})

    success_count, failed_count, failed_rows = await service.bulk_import(items)

    return MaterialImportResult(
        success_count=success_count,
        failed_count=failed_count + len(errors),
        failed_rows=failed_rows + errors,
    )
