"""BOM 文件解析服务."""

from typing import NamedTuple
from dataclasses import dataclass
from datetime import datetime
from openpyxl import load_workbook


class ParsedMaterial(NamedTuple):
    """解析后的物料行."""

    level: str
    part_number: str
    part_name: str
    version: str
    type: str
    status: str
    material: str
    supplier: str
    quantity: float
    unit: str
    comments: str


class ParsedProcess(NamedTuple):
    """解析后的工艺行."""

    op_no: str
    name: str
    work_center: str
    standard_time: float
    spec: str | None


class BOMParseResult(NamedTuple):
    """BOM 解析结果."""

    materials: list[ParsedMaterial]
    processes: list[ParsedProcess]


class ColumnMapping(NamedTuple):
    """列映射配置."""

    level: int
    part_number: int
    part_name: int
    version: int
    type: int
    status: int
    material: int
    supplier: int
    quantity: int
    unit: int
    comments: int


class BOMParser:
    """BOM 文件解析器 - 支持智能列检测."""

    # 默认列映射（原始格式）
    DEFAULT_MAPPING = ColumnMapping(
        level=0,
        part_number=1,
        part_name=2,
        version=3,
        type=4,
        status=5,
        material=6,
        supplier=7,
        quantity=8,
        unit=9,
        comments=12,
    )

    # 实际 BOM 文件列映射（基于 tests/files/bom.xlsx）
    ACTUAL_BOM_MAPPING = ColumnMapping(
        level=0,
        part_number=4,
        part_name=5,
        version=6,
        type=7,
        status=8,
        material=9,
        supplier=10,
        quantity=11,
        unit=12,
        comments=13,
    )

    def parse_excel_file(self, file_content: bytes) -> BOMParseResult:
        """解析 Excel BOM 文件（从内存）.

        Args:
            file_content: Excel 文件的字节内容

        Returns:
            BOMParseResult: 解析后的物料和工艺列表
        """
        import io

        wb = load_workbook(filename=io.BytesIO(file_content), read_only=True)

        materials: list[ParsedMaterial] = []
        processes: list[ParsedProcess] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_type = self._detect_sheet_type(ws)

            if sheet_type == "info":
                # 跳过信息类 sheet
                continue
            elif sheet_type == "material":
                # 智能检测列映射
                header_row, mapping = self._detect_column_mapping(ws)
                materials.extend(self._parse_material_sheet(ws, header_row, mapping))
            elif sheet_type == "process":
                processes.extend(self._parse_process_sheet(ws))

        return BOMParseResult(materials=materials, processes=processes)

    def _detect_sheet_type(self, worksheet) -> str:
        """检测工作表类型.

        通过关键字检测工作表包含物料还是工艺数据。
        """
        keywords_material = ["物料", "material", "bom", "item", "零件", "bill of material"]
        keywords_process = ["工艺", "process", "operation", "工序", "op"]

        # 排除的信息类 sheet
        exclude_keywords = ["product info", "产品信息", "产品名称"]

        for row in worksheet.iter_rows(min_row=1, max_row=10, values_only=True):
            if not row:
                continue
            row_text = " ".join(str(cell).lower() for cell in row if cell)

            # 先检查是否为排除的 sheet
            for kw in exclude_keywords:
                if kw in row_text:
                    return "info"

            for kw in keywords_material:
                if kw in row_text:
                    return "material"

            for kw in keywords_process:
                if kw in row_text:
                    return "process"

        return "unknown"

    def _detect_column_mapping(self, worksheet) -> tuple[int, ColumnMapping]:
        """智能检测列映射和表头行位置.

        Returns:
            (header_row, column_mapping): 表头行号和列映射
        """
        # 查找表头行（包含 "Part Number" 或 "零件号" 的行）
        header_row = 1
        for i, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            row_text = " ".join(str(cell).lower() for cell in row if cell)
            if "part number" in row_text or "零件号" in row_text:
                header_row = i
                break

        # 检测列位置
        col_map = {}
        for col_idx, cell in enumerate(worksheet[header_row]):
            if cell.value is None:
                continue
            cell_lower = str(cell.value).lower()

            if "part number" in cell_lower or "零件号" in cell_lower:
                col_map["part_number"] = col_idx
            elif "part name" in cell_lower or "零件名称" in cell_lower:
                col_map["part_name"] = col_idx
            elif "ver" in cell_lower and "version" not in cell_lower:
                col_map["version"] = col_idx
            elif "qty" in cell_lower or "quantity" in cell_lower or "数量" in cell_lower:
                col_map["quantity"] = col_idx
            elif "unit" in cell_lower or "单位" in cell_lower:
                col_map["unit"] = col_idx
            elif "comments" in cell_lower or "备注" in cell_lower or "comment" in cell_lower:
                col_map["comments"] = col_idx
            elif "material" in cell_lower and "supplier" not in cell_lower:
                col_map["material"] = col_idx
            elif "supplier" in cell_lower or "供应商" in cell_lower:
                col_map["supplier"] = col_idx
            elif "typ" in cell_lower and "part" not in cell_lower:
                col_map["type"] = col_idx
            elif "st" in cell_lower and len(cell_lower) <= 3:
                col_map["status"] = col_idx
            elif "level" in cell_lower or "层级" in cell_lower:
                col_map["level"] = col_idx

        # 如果检测到足够的关键列，使用检测到的映射
        if len(col_map) >= 5:
            return header_row, ColumnMapping(
                level=col_map.get("level", 0),
                part_number=col_map.get("part_number", 1),
                part_name=col_map.get("part_name", 2),
                version=col_map.get("version", 3),
                type=col_map.get("type", 4),
                status=col_map.get("status", 5),
                material=col_map.get("material", 6),
                supplier=col_map.get("supplier", 7),
                quantity=col_map.get("quantity", 8),
                unit=col_map.get("unit", 9),
                comments=col_map.get("comments", 10),
            )

        # 否则使用默认映射
        return header_row, self.DEFAULT_MAPPING

    def _parse_material_sheet(
        self, worksheet, header_row: int, mapping: ColumnMapping
    ) -> list[ParsedMaterial]:
        """解析物料工作表.

        Args:
            worksheet: 工作表对象
            header_row: 表头行号
            mapping: 列映射配置
        """
        materials = []

        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or len(row) <= max(mapping.part_number, mapping.quantity):
                continue

            # 获取零件号
            part_number_cell = row[mapping.part_number] if mapping.part_number < len(row) else None
            if not part_number_cell:
                continue

            part_number = str(part_number_cell).strip()
            # 跳过空值或表头行
            if not part_number or part_number.lower() in ["", "none", "part number", "零件号"]:
                continue

            # 获取数量
            quantity = 0
            if mapping.quantity < len(row) and row[mapping.quantity] is not None:
                try:
                    quantity = float(row[mapping.quantity])
                except (ValueError, TypeError):
                    quantity = 0

            materials.append(
                ParsedMaterial(
                    level=str(row[mapping.level] or "") if mapping.level < len(row) else "1",
                    part_number=part_number,
                    part_name=str(row[mapping.part_name] or "") if mapping.part_name < len(row) else "",
                    version=str(row[mapping.version] or "1.0") if mapping.version < len(row) else "1.0",
                    type=str(row[mapping.type] or "I") if mapping.type < len(row) else "I",
                    status=str(row[mapping.status] or "N") if mapping.status < len(row) else "N",
                    material=str(row[mapping.material] or "") if mapping.material < len(row) else "",
                    supplier=str(row[mapping.supplier] or "") if mapping.supplier < len(row) else "",
                    quantity=quantity,
                    unit=str(row[mapping.unit] or "PC") if mapping.unit < len(row) else "PC",
                    comments=str(row[mapping.comments] or "") if mapping.comments < len(row) else "",
                )
            )

        return materials

    def _parse_process_sheet(self, worksheet) -> list[ParsedProcess]:
        """解析工艺工作表.

        默认列映射:
        - Col 0: Op No (工序号)
        - Col 1: Name (工序名称)
        - Col 2: Work Center (工作中心)
        - Col 3: Standard Time (标准工时)
        - Col 4: Spec (规格说明)
        """
        processes = []

        # 查找表头行
        header_row = 1
        for i, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            row_text = " ".join(str(cell).lower() for cell in row if cell)
            if "op no" in row_text or "工序号" in row_text or "operation" in row_text:
                header_row = i
                break

        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[0]:
                continue

            op_no = str(row[0]).strip()
            if not op_no or op_no.lower() in ["", "none", "op no", "工序号"]:
                continue

            # 安全获取各列数据，处理行数据不完整的情况
            name = str(row[1]) if len(row) > 1 and row[1] else ""
            work_center = str(row[2]) if len(row) > 2 and row[2] else ""

            # 标准工时可能是分钟或小时，存储为分钟数
            standard_time = 0
            if len(row) > 3 and row[3] is not None:
                try:
                    standard_time = float(row[3])
                except (ValueError, TypeError):
                    standard_time = 0

            spec = str(row[4]) if len(row) > 4 and row[4] else None

            processes.append(
                ParsedProcess(
                    op_no=op_no,
                    name=name,
                    work_center=work_center,
                    standard_time=standard_time,
                    spec=spec,
                )
            )

        return processes


# ==================== 多产品 BOM 解析 ====================

@dataclass
class ProductInfo:
    """产品元数据（从 sheet 顶部提取）."""
    product_code: str          # 从 sheet 名称获取
    product_name: str | None   # 从 "Product Name" 字段提取
    product_number: str | None # 从 "Product Number" 字段提取
    product_version: str       # 默认 "01"
    customer_version: str      # 默认 "01"
    customer_number: str | None # 从 "Customer Number" 字段提取
    issue_date: datetime | None # 从 "Issue Date" 字段提取
    material_count: int = 0
    process_count: int = 0


@dataclass
class ProductBOMResult:
    """单个产品的 BOM 解析结果."""
    product_info: ProductInfo
    materials: list[ParsedMaterial]
    processes: list[ParsedProcess]


@dataclass
class MultiProductBOMParseResult:
    """多产品 BOM 解析结果."""
    products: list[ProductBOMResult]
    total_products: int
    total_materials: int
    parse_warnings: list[str]


class MultiProductBOMParser:
    """多产品 BOM 解析器.

    用于解析包含多个产品的 BOM 文件，每个 sheet 代表一个产品。
    """

    def __init__(self):
        """初始化解析器."""
        # 复用现有的 BOMParser 方法
        self._base_parser = BOMParser()

    def parse_excel_file(self, file_content: bytes) -> MultiProductBOMParseResult:
        """解析多产品 Excel BOM 文件.

        Args:
            file_content: Excel 文件的字节内容

        Returns:
            MultiProductBOMParseResult: 解析后的多产品结果
        """
        import io

        wb = load_workbook(filename=io.BytesIO(file_content), read_only=True)
        results = []
        warnings = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 检测 sheet 类型，跳过非物料 sheet
            sheet_type = self._base_parser._detect_sheet_type(ws)
            if sheet_type != "material":
                continue

            # 1. 提取产品元数据
            product_info = self._extract_product_metadata(ws, sheet_name)

            # 2. 解析物料列表
            materials = self._parse_materials_for_product(ws)
            product_info.material_count = len(materials)

            # 3. 工艺路线暂为空
            processes = []
            product_info.process_count = 0

            results.append(ProductBOMResult(
                product_info=product_info,
                materials=materials,
                processes=processes
            ))

        return MultiProductBOMParseResult(
            products=results,
            total_products=len(results),
            total_materials=sum(r.product_info.material_count for r in results),
            parse_warnings=warnings
        )

    def _extract_product_metadata(self, worksheet, sheet_name: str) -> ProductInfo:
        """从 sheet 顶部提取产品元数据.

        Sheet 实际结构（基于 W04_BOM_.xlsx 分析）:
        Row 2 (index 1): Product Name (col 5), Product Version (col 11) -> 值在 col 13
        Row 3 (index 2): Product Number (col 5) -> 值在 col 6, Customer Version (col 11) -> 值在 col 13
        Row 4 (index 3): Customer Number (col 5) -> 值在 col 6, Issue Date (col 11) -> 值在 col 13
        """
        product_code = sheet_name.strip()
        product_name = None
        product_number = None
        product_version = "01"
        customer_version = "01"
        customer_number = None
        issue_date = None

        # 检查前 10 行，寻找包含关键标签的行
        for row_idx in range(1, min(11, worksheet.max_row)):
            row = worksheet[row_idx]
            row_values = [cell.value for cell in row]

            for col_idx, cell_value in enumerate(row_values):
                if cell_value and isinstance(cell_value, str):
                    key = cell_value.rstrip(':').strip().lower()

                    # Product Name: 标签在 col 5, 值可能在 col 6 或 7
                    if 'product name' in key and col_idx == 5:
                        # 尝试 col 6, 7 找值
                        for val_col in [6, 7]:
                            if len(row_values) > val_col and row_values[val_col]:
                                val = str(row_values[val_col]).strip()
                                if val and val != 'Product Name':
                                    product_name = val
                                    break

                    # Product Number: 标签在 col 5, 值在 col 6
                    elif 'product number' in key and col_idx == 5:
                        if len(row_values) > 6 and row_values[6]:
                            val = str(row_values[6]).strip()
                            if val and val != 'Product Number':
                                product_number = val

                    # Product Version: 标签在 col 11, 值在 col 13
                    elif 'product version' in key and col_idx == 11:
                        if len(row_values) > 13 and row_values[13]:
                            val = str(row_values[13]).strip()
                            if val:
                                product_version = val

                    # Customer Version: 标签在 col 11, 值在 col 13
                    elif 'customer version' in key and col_idx == 11:
                        if len(row_values) > 13 and row_values[13]:
                            val = str(row_values[13]).strip()
                            if val:
                                customer_version = val

                    # Customer Number: 标签在 col 5, 值在 col 6
                    elif 'customer number' in key and col_idx == 5:
                        if len(row_values) > 6 and row_values[6]:
                            val = str(row_values[6]).strip()
                            if val and val != 'Customer Number':
                                customer_number = val

                    # Issue Date: 标签在 col 11, 值在 col 13
                    elif 'issue date' in key and col_idx == 11:
                        if len(row_values) > 13 and row_values[13]:
                            val = row_values[13]
                            # 处理 Excel 日期格式
                            if isinstance(val, datetime):
                                issue_date = val.isoformat()
                            else:
                                issue_date = str(val)

        return ProductInfo(
            product_code=product_code,
            product_name=product_name,
            product_number=product_number,
            product_version=product_version,
            customer_version=customer_version,
            customer_number=customer_number,
            issue_date=self._parse_date(issue_date),
        )

    def _parse_date(self, date_input: str | datetime | None) -> datetime | None:
        """解析日期（支持字符串、datetime 对象）."""
        if not date_input:
            return None
        try:
            # 如果已经是 datetime 对象，直接返回
            if isinstance(date_input, datetime):
                return date_input
            # 处理字符串日期
            if isinstance(date_input, str):
                # 处理 Excel 日期序列号
                if date_input.isdigit():
                    return datetime.fromordinal(int(date_input) + 693594)  # Excel epoch 1900-01-01
                # 处理 ISO 格式日期
                return datetime.fromisoformat(date_input)
        except (ValueError, TypeError):
            return None

    def _parse_materials_for_product(self, worksheet) -> list[ParsedMaterial]:
        """为单个产品解析物料.

        Args:
            worksheet: 工作表对象

        Returns:
            list[ParsedMaterial]: 物料列表
        """
        # 使用现有的智能列检测
        header_row, mapping = self._base_parser._detect_column_mapping(worksheet)
        return self._base_parser._parse_material_sheet(worksheet, header_row, mapping)
